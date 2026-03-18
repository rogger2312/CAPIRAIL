import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse

from models import StepByStepConvertInput
from services.auth import verify_token
from services.csv_generator import (
    cases_to_csv,
    cases_to_xml,
    build_testrail_cases,
    cases_to_testrail_steps_csv,
)

router = APIRouter()


@router.post("/api/download/csv")
def download_csv(payload: dict, _: None = Depends(verify_token)):
    cases = payload.get("test_cases", [])
    if not cases:
        raise HTTPException(status_code=400, detail="No hay casos de prueba para exportar")

    csv_content = cases_to_csv(cases)
    return StreamingResponse(
        io.BytesIO(csv_content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=test_cases_testrail.csv"},
    )


@router.post("/api/download/xml")
def download_xml(payload: dict, _: None = Depends(verify_token)):
    """Genera XML en formato oficial de TestRail para importar directamente."""
    cases = payload.get("test_cases", [])
    section_path = payload.get("section_path", "").strip()
    if not cases:
        raise HTTPException(status_code=400, detail="No hay casos de prueba para exportar")

    xml_content = cases_to_xml(cases, section_path=section_path)
    return StreamingResponse(
        io.BytesIO(xml_content.encode("utf-8")),
        media_type="application/xml",
        headers={"Content-Disposition": "attachment; filename=test_cases_testrail.xml"},
    )


@router.post("/api/download/titles-xlsx")
def download_titles_xlsx(payload: dict, _: None = Depends(verify_token)):
    """Descarga los títulos generados como archivo Excel (.xlsx)."""
    titles = payload.get("titles", [])
    if not titles:
        raise HTTPException(status_code=400, detail="No hay títulos para exportar")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Case Titles"

    # Cabecera
    ws.append(["#", "Test Case Title"])
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 90

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="0F3460")
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Filas
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, title in enumerate(titles, 1):
        ws.append([i, str(title)])
        ws.cell(row=i + 1, column=2).alignment = wrap

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_case_titles.xlsx"},
    )


@router.post("/api/download/steps-csv")
def download_steps_csv(payload: StepByStepConvertInput, _: None = Depends(verify_token)):
    """Convert parsed rows to TestRail step-by-step CSV and download."""
    if not payload.rows:
        raise HTTPException(status_code=400, detail="No hay datos para exportar")
    cases = build_testrail_cases(payload.rows, payload.col_map, payload.defaults)
    if not cases:
        raise HTTPException(status_code=400, detail="No se pudieron extraer casos del archivo")
    csv_content = cases_to_testrail_steps_csv(cases)
    return StreamingResponse(
        io.BytesIO(csv_content.encode('utf-8-sig')),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=testrail_import.csv'},
    )

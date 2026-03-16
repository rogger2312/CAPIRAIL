import os
import csv
import json
import io
import base64
import openpyxl
import xml.etree.ElementTree as ET
from xml.dom import minidom
from collections import defaultdict
from typing import List, Optional
from fastapi import FastAPI, HTTPException, UploadFile, File, Depends, Header
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from google import genai
from google.genai import types as genai_types
from dotenv import load_dotenv

load_dotenv()

APP_PASSWORD = os.getenv("APP_PASSWORD", "")

app = FastAPI(title="TestRail BDD Generator")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")


# ─── Modelos de entrada ────────────────────────────────────────────────────────

class JiraInput(BaseModel):
    description: str
    acceptance_criteria: str
    jira_ticket: Optional[str] = ""
    images: List[str] = []   # base64 data-URLs (máx 3)

class StepInput(BaseModel):
    step: str
    validation: str

class TestRailInput(BaseModel):
    preconditions: str
    test_data: Optional[str] = ""
    steps: List[StepInput]
    title_hint: Optional[str] = ""
    images: List[str] = []   # base64 data-URLs (máx 3)

class TitleGeneratorInput(BaseModel):
    test_cases: str          # descripciones de casos, una por línea o texto libre
    images: List[str] = []   # base64 data-URLs opcionales (máx 3)

class BatchItem(BaseModel):
    title: str
    section: str
    refs: str
    description: Optional[str] = ""

class BatchGenerateInput(BaseModel):
    items: List[BatchItem]
    type_override: str = "Functional"
    priority_override: str = "3 - Normal"
    extra_fields: dict = {}


# ─── Modelo de salida (caso de prueba) ────────────────────────────────────────

class TestCase(BaseModel):
    title: str
    section: str
    template: str
    type: str
    priority: str
    estimate: str
    refs: str
    custom_preconds: str
    custom_bdd_scenarios: str


# ─── Cliente Gemini ───────────────────────────────────────────────────────────

def get_gemini_client():
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY no configurada en el archivo .env")
    return genai.Client(api_key=api_key, http_options={"api_version": "v1beta"})


# ─── Prompts ──────────────────────────────────────────────────────────────────

TITLE_SYSTEM_PROMPT = """You are a QA expert specializing in test case naming conventions.
Your task is to generate clear, professional test case titles in English.

STRICT RULES:
- Titles must NOT contain the words "verify", "validate", or "confirm" (in any form or capitalization).
- Titles must NOT include sensitive data (passwords, tokens, personal information, credentials).
- Each title must clearly indicate the specific purpose of the test, describing what is being checked or evaluated.
- All titles must be in English.
- Titles should be concise but descriptive (max 120 characters).
- Use action verbs such as: "Check", "Ensure", "Test", "Assert", "Evaluate", "Examine", "Display", "Prevent", "Allow", "Handle", "Return", "Submit", "Navigate", etc.
- If images are provided, analyze them to understand the feature being tested and use that context.

Respond ONLY with a valid JSON array of strings (the titles), no explanations, no markdown code blocks.
Example: ["Check that the login page displays an error for incorrect credentials", "Ensure the dashboard loads after a successful authentication"]
"""

SYSTEM_PROMPT = """Eres un experto en QA y testing de software especializado en metodología BDD (Behavior Driven Development).
Tu tarea es generar casos de prueba en formato Gherkin listos para importar en TestRail usando el template "Behaviour Driven Development".

REGLAS IMPORTANTES:
- Determina la cantidad óptima de casos de prueba según la complejidad del requerimiento (mínimo 3, máximo 15).
- Cada caso debe cubrir un escenario distinto (happy path, casos borde, errores, seguridad, etc.).
- Si se incluyen imágenes, analízalas detalladamente y úsalas como contexto visual para generar mejores escenarios.
- Si se incluyen fragmentos de código o JSON, analízalos para entender la estructura de datos y generar escenarios precisos.
- Responde ÚNICAMENTE con un array JSON válido, sin explicaciones adicionales, sin bloques de código markdown.
- El JSON debe seguir exactamente el esquema indicado.
- TODOS los textos del JSON (títulos, precondiciones, pasos, resultados esperados) deben estar en INGLÉS.

FORMATO DE PRIORIDADES (usar exactamente estos valores):
- "1 - Must Test"
- "2 - Run Before Release"
- "3 - Normal"
- "4 - Nice to Have"

FORMATO custom_bdd_scenarios: Escenario Gherkin completo con Feature, Background (si aplica) y Scenario.
Usa \\n para saltos de línea y dos espacios de indentación.
Ejemplo:
  Feature: User Login\\n  Background:\\n    Given the user is on the login page\\n  Scenario: Successful login\\n    When the user enters valid credentials\\n    Then the user is redirected to the dashboard
"""

def prompt_from_jira(data: JiraInput) -> str:
    return f"""Generate BDD test cases from the following Jira user story.

USER STORY:
{data.description}

ACCEPTANCE CRITERIA:
{data.acceptance_criteria}

REFERENCE TICKET: {data.jira_ticket or "N/A"}

Return a JSON array where each object has EXACTLY these keys:
- "title": Descriptive test case title (max 120 chars)
- "section": Section/suite name (e.g. "Login", "Checkout", etc.)
- "template": Always use "Behaviour Driven Development"
- "type": Test type ("Functional", "Integration", "Security", "Regression", "Acceptance")
- "priority": Exactly one of: "1 - Must Test", "2 - Run Before Release", "3 - Normal", "4 - Nice to Have"
- "estimate": Time estimate in minutes (e.g. "3m", "5m", "10m")
- "refs": Reference to Jira ticket (use the ticket provided or leave empty)
- "custom_preconds": Preconditions required to run the test case
- "custom_bdd_scenarios": Complete Gherkin scenario (Feature + optional Background + Scenario) with newlines as \\n
"""

def prompt_from_testrail(data: TestRailInput) -> str:
    steps_text = "\n".join(
        f"Step {i+1}: {s.step}\n  Validation: {s.validation}"
        for i, s in enumerate(data.steps)
    )
    return f"""Transform the following manual test case into improved BDD test cases.

TITLE / CONTEXT: {data.title_hint or "Not specified"}

PRECONDITIONS:
{data.preconditions}

TEST DATA:
{data.test_data or "Not specified"}

STEPS AND VALIDATIONS:
{steps_text}

Analyze these steps and generate multiple BDD scenarios covering:
1. The main flow (happy path)
2. Variations with test data
3. Relevant error or edge cases

Return a JSON array where each object has EXACTLY these keys:
- "title": Descriptive test case title (max 120 chars)
- "section": Section/suite name inferred from context
- "template": Always use "Behaviour Driven Development"
- "type": Test type ("Functional", "Integration", "Security", "Regression", "Acceptance")
- "priority": Exactly one of: "1 - Must Test", "2 - Run Before Release", "3 - Normal", "4 - Nice to Have"
- "estimate": Time estimate in minutes (e.g. "3m", "5m", "10m")
- "refs": Leave empty ("")
- "custom_preconds": Required preconditions
- "custom_bdd_scenarios": Complete Gherkin scenario (Feature + optional Background + Scenario) with newlines as \\n
"""

def prompt_from_batch_item(item: BatchItem) -> str:
    desc_block = f"\nDESCRIPTION / CONTEXT:\n{item.description}" if item.description else ""
    return f"""Generate BDD test cases for the following requirement.

TITLE / CONTEXT: {item.title}
SECTION: {item.section or "Not specified"}
REFERENCE: {item.refs or "N/A"}{desc_block}

Return a JSON array where each object has EXACTLY these keys:
- "title": Descriptive BDD test case title based on the context (max 120 chars)
- "section": Use "{item.section or 'General'}" as the section name
- "template": Always use "Behaviour Driven Development"
- "type": Test type ("Functional", "Integration", "Security", "Regression", "Acceptance")
- "priority": Exactly one of: "1 - Must Test", "2 - Run Before Release", "3 - Normal", "4 - Nice to Have"
- "estimate": Time estimate in minutes (e.g. "3m", "5m", "10m")
- "refs": Use "{item.refs}" as reference
- "custom_preconds": Preconditions required to run the test case
- "custom_bdd_scenarios": Complete Gherkin scenario (Feature + optional Background + Scenario) with newlines as \\n
"""


# ─── Función de llamada a Gemini ─────────────────────────────────────────────

def call_gemini(user_prompt: str, images: List[str] = []) -> List[dict]:
    try:
        client = get_gemini_client()

        # Construir contenido multimodal si hay imágenes
        if images:
            parts: list = [genai_types.Part.from_text(text=user_prompt)]
            for img_b64 in images[:3]:
                if ',' in img_b64:
                    header, data = img_b64.split(',', 1)
                    mime = header.split(';')[0].split(':')[1]
                else:
                    data = img_b64
                    mime = 'image/jpeg'
                parts.append(genai_types.Part.from_bytes(
                    data=base64.b64decode(data),
                    mime_type=mime,
                ))
            contents = parts
        else:
            contents = user_prompt

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            config=genai_types.GenerateContentConfig(
                system_instruction=SYSTEM_PROMPT,
            ),
            contents=contents,
        )
        text_content = response.text
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al llamar a Gemini: {type(e).__name__}: {str(e)}")

    if not text_content:
        raise HTTPException(status_code=500, detail="El modelo no devolvió contenido de texto")

    text_content = text_content.strip()
    if text_content.startswith("```"):
        lines = text_content.split("\n")
        text_content = "\n".join(lines[1:])
        if text_content.endswith("```"):
            text_content = text_content[:-3].strip()

    try:
        test_cases = json.loads(text_content)
    except json.JSONDecodeError as e:
        raise HTTPException(
            status_code=500,
            detail=f"El modelo devolvió JSON inválido: {str(e)}\n\nRespuesta: {text_content[:500]}"
        )

    if not isinstance(test_cases, list):
        raise HTTPException(status_code=500, detail="El modelo no devolvió un array de casos")

    return test_cases

call_claude = call_gemini  # alias


# ─── Parser de XML ────────────────────────────────────────────────────────────

def parse_xml_content(content: str) -> List[dict]:
    """
    Soporta tres formatos:
      1. TestRail XML export: <sections><section><cases><case>
      2. Jira RSS export:     <rss><channel><item>
      3. Formato simple:      <cases><case> / <test-cases><test-case> / etc.
    """
    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        raise HTTPException(status_code=400, detail=f"XML inválido: {str(e)}")

    items = []

    # ── 1. TestRail XML export: <suite><sections>... o <sections>... ──────────
    # Detectar por root tag o presencia de <section>/<cases>/<case>
    tr_sections = root.findall('.//section')
    has_tr_cases = root.find('.//cases/case') is not None
    if (root.tag in ('suite', 'sections') or has_tr_cases) and tr_sections:
        for section_el in root.findall('.//section'):
            section_name = (section_el.findtext('name') or '').strip()
            # Solo casos directos de esta section (no de subsecciones)
            cases_el = section_el.find('cases')
            if cases_el is None:
                continue
            for case_el in cases_el.findall('case'):
                title = (case_el.findtext('title') or '').strip()
                refs  = (case_el.findtext('refs')  or '').strip()
                # Buscar descripción en custom fields
                custom_el = case_el.find('custom')
                description = ''
                if custom_el is not None:
                    description = (
                        custom_el.findtext('bdd_scenarios') or
                        custom_el.findtext('preconds') or
                        custom_el.findtext('steps') or ''
                    ).strip()
                if title:
                    items.append({
                        'title':       title,
                        'section':     section_name,
                        'refs':        refs,
                        'description': description,
                    })
        if items:
            return items

    # ── 2. Jira RSS export: <rss><channel><item> ──────────────────────────────
    jira_items = root.findall('.//item')
    if jira_items:
        for item in jira_items:
            summary   = (item.findtext('summary') or '').strip()
            raw_title = (item.findtext('title')   or '').strip()
            if raw_title.startswith('[') and '] ' in raw_title:
                raw_title = raw_title.split('] ', 1)[1]
            title = summary or raw_title

            key         = (item.findtext('key')         or '').strip()
            description = (item.findtext('description') or '').strip()

            component_el = item.find('component')
            section = (component_el.text if component_el is not None else '').strip()

            if title:
                items.append({
                    'title':       title,
                    'section':     section,
                    'refs':        key,
                    'description': description,
                })
        return items

    # ── 3. Formato simple ─────────────────────────────────────────────────────
    for tag in ['case', 'test-case', 'testcase', 'story', 'issue', 'requirement']:
        elements = root.findall(f'.//{tag}')
        if elements:
            for el in elements:
                title = (
                    el.findtext('title') or
                    el.findtext('summary') or
                    el.get('title') or ''
                ).strip()
                section = (el.findtext('section') or el.findtext('component') or '').strip()
                refs    = (
                    el.findtext('refs') or
                    el.findtext('references') or
                    el.findtext('key') or ''
                ).strip()
                description = (el.findtext('description') or el.findtext('body') or '').strip()
                if title:
                    items.append({
                        'title':       title,
                        'section':     section,
                        'refs':        refs,
                        'description': description,
                    })
            return items

    return items


# ─── Generador de CSV ─────────────────────────────────────────────────────────

BASE_CSV_COLUMNS = [
    "title", "section", "template", "type", "priority", "estimate",
    "refs", "custom_preconds", "custom_bdd_scenarios"
]

def cases_to_csv(cases: List[dict]) -> str:
    extra_keys = []
    for case in cases:
        for k in case.keys():
            if k not in BASE_CSV_COLUMNS and k not in extra_keys:
                extra_keys.append(k)
    extra_keys.sort()
    columns = BASE_CSV_COLUMNS + extra_keys

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=columns,
        quoting=csv.QUOTE_ALL,
        lineterminator="\r\n",
        extrasaction='ignore',
    )
    writer.writeheader()

    for case in cases:
        row = {col: case.get(col, '') for col in columns}
        row['template'] = case.get('template', 'Behaviour Driven Development')
        row['type']     = case.get('type', 'Functional')
        row['priority'] = case.get('priority', '3 - Normal')
        row['estimate'] = case.get('estimate', '5m')
        writer.writerow(row)

    return output.getvalue()


# ─── Generador de XML (formato TestRail) ─────────────────────────────────────

# Campos estándar de TestRail a nivel de <case> (fuera de <custom>)
_TR_STANDARD = {'title', 'section', 'template', 'type', 'priority', 'estimate', 'refs'}

def _priority_id(priority_str: str) -> str:
    """Convierte '2 - Run Before Release' → '2' (TestRail espera ID numérico en XML)."""
    part = priority_str.strip().split(' ')[0]
    return part if part.isdigit() else '3'


def cases_to_xml(cases: List[dict]) -> str:
    """Genera XML en formato TestRail probado:
       - Raíz <sections> (import en suite existente)
       - <references> (no <refs>)
       - <priority> como número
       - CDATA para contenido Gherkin
    """
    # Agrupar por sección manteniendo orden de aparición
    sections: dict = {}
    for case in cases:
        sec = case.get('section', 'General') or 'General'
        sections.setdefault(sec, []).append(case)

    root = ET.Element('sections')

    for section_name, section_cases in sections.items():
        sec_el   = ET.SubElement(root, 'section')
        ET.SubElement(sec_el, 'name').text = section_name
        cases_el = ET.SubElement(sec_el, 'cases')

        for case in section_cases:
            case_el = ET.SubElement(cases_el, 'case')

            ET.SubElement(case_el, 'title').text    = case.get('title', '')
            ET.SubElement(case_el, 'type').text     = case.get('type', 'Functional')
            ET.SubElement(case_el, 'priority').text = _priority_id(case.get('priority', '3 - Normal'))
            refs = case.get('refs', '').strip()
            # Siempre incluir <references> (campo requerido en TestRail).
            # Si no hay ticket, usar "-" para evitar que TestRail lo valide contra Jira.
            ET.SubElement(case_el, 'references').text = refs if refs else '-'

            estimate = case.get('estimate', '')
            if estimate:
                ET.SubElement(case_el, 'estimate').text = estimate

            # Bloque <custom>
            custom_el = ET.SubElement(case_el, 'custom')

            preconds = case.get('custom_preconds', '')
            if preconds:
                ET.SubElement(custom_el, 'preconds').text = preconds

            bdd = case.get('custom_bdd_scenarios', '')
            if bdd:
                ET.SubElement(custom_el, 'bdd_scenarios').text = bdd  # → CDATA en minidom

            # Campos extra (implementation, domain, etc.)
            skip = _TR_STANDARD | {'custom_preconds', 'custom_bdd_scenarios'}
            for key, value in case.items():
                if key not in skip and value:
                    field_name = key[7:] if key.startswith('custom_') else key
                    ET.SubElement(custom_el, field_name).text = str(value)

    # Convertir a DOM de minidom para aplicar CDATA en bdd_scenarios y pretty-print
    raw = ET.tostring(root, encoding='unicode')
    dom = minidom.parseString(raw)

    # Envolver contenido de <bdd_scenarios> y <preconds> en CDATA (sin whitespace extra)
    for tag in ('bdd_scenarios', 'preconds'):
        for el in dom.getElementsByTagName(tag):
            if el.firstChild and el.firstChild.nodeType == el.firstChild.TEXT_NODE:
                text = el.firstChild.data.strip()
                el.removeChild(el.firstChild)
                el.appendChild(dom.createCDATASection(text))

    pretty = dom.toprettyxml(indent='  ', encoding='UTF-8')
    return pretty.decode('UTF-8') if isinstance(pretty, bytes) else pretty


# ─── Auth ─────────────────────────────────────────────────────────────────────

def verify_token(authorization: Optional[str] = Header(None)):
    if not APP_PASSWORD:
        return  # sin contraseña configurada, acceso libre
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="No autorizado")
    if authorization.split(" ", 1)[1] != APP_PASSWORD:
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")


# ─── Endpoints ────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return FileResponse("static/index.html")


@app.post("/api/login")
def login(payload: dict):
    if not APP_PASSWORD:
        return {"token": "open"}
    if payload.get("password") == APP_PASSWORD:
        return {"token": APP_PASSWORD}
    raise HTTPException(status_code=401, detail="Contraseña incorrecta")


@app.post("/api/generate/jira")
def generate_from_jira(data: JiraInput, _: None = Depends(verify_token)):
    prompt = prompt_from_jira(data)
    cases  = call_gemini(prompt, images=data.images)
    return {"test_cases": cases, "count": len(cases)}


@app.post("/api/generate/testrail")
def generate_from_testrail(data: TestRailInput, _: None = Depends(verify_token)):
    if not data.steps:
        raise HTTPException(status_code=400, detail="Debes proporcionar al menos un paso")
    prompt = prompt_from_testrail(data)
    cases  = call_gemini(prompt, images=data.images)
    return {"test_cases": cases, "count": len(cases)}


@app.post("/api/parse-xml")
async def parse_xml_endpoint(file: UploadFile = File(...), _: None = Depends(verify_token)):
    """Parsea XML de TestRail export, Jira RSS o formato simple."""
    content = await file.read()
    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        text = content.decode('latin-1')

    items = parse_xml_content(text)
    if not items:
        raise HTTPException(
            status_code=400,
            detail=(
                "No se encontraron items en el XML. "
                "Formatos soportados: TestRail XML export (<sections>), "
                "Jira RSS export (<rss>), o formato simple (<cases><case>)."
            )
        )
    return {"items": items, "count": len(items)}


@app.post("/api/generate/batch")
def generate_batch(config: BatchGenerateInput, _: None = Depends(verify_token)):
    if not config.items:
        raise HTTPException(status_code=400, detail="No hay items para procesar")

    all_cases = []
    errors    = []

    for item in config.items:
        try:
            prompt = prompt_from_batch_item(item)
            cases  = call_gemini(prompt)
            for case in cases:
                case['type']     = config.type_override
                case['priority'] = config.priority_override
                if item.refs:
                    case['refs'] = item.refs
                for key, value in config.extra_fields.items():
                    if value:
                        case[key] = value
            all_cases.extend(cases)
        except Exception as e:
            errors.append({"item": item.title, "error": str(e)})

    return {"test_cases": all_cases, "count": len(all_cases), "errors": errors}


def parse_testrail_xml_full(content: str) -> List[dict]:
    """Parsea XML de TestRail completo: extrae TODOS los campos de cada caso."""
    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        raise HTTPException(status_code=400, detail=f"XML inválido: {str(e)}")

    cases = []
    for section_el in root.findall('.//section'):
        section_name = (section_el.findtext('name') or 'General').strip()
        cases_el = section_el.find('cases')
        if cases_el is None:
            continue
        for case_el in cases_el.findall('case'):
            case = {
                'title':    (case_el.findtext('title')    or '').strip(),
                'section':  section_name,
                'template': (case_el.findtext('template') or 'Behaviour Driven Development').strip(),
                'type':     (case_el.findtext('type')     or 'Functional').strip(),
                'priority': (case_el.findtext('priority') or '3 - Normal').strip(),
                'estimate': (case_el.findtext('estimate') or '').strip(),
                'refs':     (case_el.findtext('refs') or case_el.findtext('references') or '').strip(),
                'custom_preconds':      '',
                'custom_bdd_scenarios': '',
            }
            custom_el = case_el.find('custom')
            if custom_el is not None:
                case['custom_preconds']      = (custom_el.findtext('preconds')      or '').strip()
                case['custom_bdd_scenarios'] = (custom_el.findtext('bdd_scenarios') or '').strip()
                # Campos custom extra (implementation, domain, automation_review, etc.)
                known = {'preconds', 'bdd_scenarios'}
                for child in custom_el:
                    if child.tag not in known:
                        case[f'custom_{child.tag}'] = (child.text or '').strip()
            if case['title']:
                cases.append(case)

    if not cases:
        raise HTTPException(status_code=400, detail="No se encontraron casos en el XML. Verifica que sea un XML exportado por esta herramienta o por TestRail.")

    return cases


@app.post("/api/load-xml")
async def load_xml_cases(file: UploadFile = File(...), _: None = Depends(verify_token)):
    """Carga un XML de TestRail y devuelve todos los casos con todos sus campos."""
    content = await file.read()
    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        text = content.decode('latin-1')
    cases = parse_testrail_xml_full(text)
    return {"cases": cases, "count": len(cases)}


@app.post("/api/download/csv")
def download_csv(payload: dict, _: None = Depends(verify_token)):
    cases = payload.get("test_cases", [])
    if not cases:
        raise HTTPException(status_code=400, detail="No hay casos de prueba para exportar")

    csv_content = cases_to_csv(cases)
    return StreamingResponse(
        io.BytesIO(csv_content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=test_cases_testrail.csv"},
    )


@app.post("/api/download/xml")
def download_xml(payload: dict, _: None = Depends(verify_token)):
    """Genera XML en formato oficial de TestRail para importar directamente."""
    cases = payload.get("test_cases", [])
    if not cases:
        raise HTTPException(status_code=400, detail="No hay casos de prueba para exportar")

    xml_content = cases_to_xml(cases)
    return StreamingResponse(
        io.BytesIO(xml_content.encode("utf-8")),
        media_type="application/xml",
        headers={"Content-Disposition": "attachment; filename=test_cases_testrail.xml"},
    )


@app.post("/api/generate/titles")
def generate_titles(data: TitleGeneratorInput, _: None = Depends(verify_token)):
    """Genera títulos de casos de prueba siguiendo convenciones estrictas de naming."""
    user_prompt = f"""Generate test case titles for the following test cases or feature description:

{data.test_cases}

Return a JSON array of strings where each string is a title for one test case.
Generate as many titles as needed to cover the described scenarios thoroughly."""

    try:
        client = get_gemini_client()

        if data.images:
            parts: list = [genai_types.Part.from_text(text=user_prompt)]
            for img_b64 in data.images[:3]:
                if ',' in img_b64:
                    header, d = img_b64.split(',', 1)
                    mime = header.split(';')[0].split(':')[1]
                else:
                    d, mime = img_b64, 'image/jpeg'
                parts.append(genai_types.Part.from_bytes(data=base64.b64decode(d), mime_type=mime))
            contents = parts
        else:
            contents = user_prompt

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            config=genai_types.GenerateContentConfig(system_instruction=TITLE_SYSTEM_PROMPT),
            contents=contents,
        )
        text_content = (response.text or '').strip()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al llamar a Gemini: {type(e).__name__}: {str(e)}")

    if text_content.startswith("```"):
        lines = text_content.split("\n")
        text_content = "\n".join(lines[1:])
        if text_content.endswith("```"):
            text_content = text_content[:-3].strip()

    try:
        titles = json.loads(text_content)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"El modelo devolvió JSON inválido: {str(e)}\n\nRespuesta: {text_content[:500]}")

    if not isinstance(titles, list):
        raise HTTPException(status_code=500, detail="El modelo no devolvió una lista de títulos")

    return {"titles": [str(t) for t in titles], "count": len(titles)}


@app.post("/api/download/titles-xlsx")
def download_titles_xlsx(payload: dict, _: None = Depends(verify_token)):
    """Descarga los títulos generados como archivo Excel (.xlsx)."""
    titles = payload.get("titles", [])
    if not titles:
        raise HTTPException(status_code=400, detail="No hay títulos para exportar")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Case Titles"

    # Cabecera
    ws.append(["#", "Test Case Title"])
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 90

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="0F3460")
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Filas
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, title in enumerate(titles, 1):
        ws.append([i, str(title)])
        ws.cell(row=i + 1, column=2).alignment = wrap

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_case_titles.xlsx"},
    )


@app.get("/api/models")
def list_models(_: None = Depends(verify_token)):
    try:
        client  = get_gemini_client()
        models  = [m.name for m in client.models.list()]
        return {"models": models}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
